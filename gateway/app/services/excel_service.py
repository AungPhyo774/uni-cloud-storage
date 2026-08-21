from openpyxl import load_workbook
import io

REQUIRED_STUDENT_COLUMNS = {
    "full_name",
    "email",
    "class_year"
}


def read_student_excel(
    file_content: bytes
):

    workbook = load_workbook(
        filename=io.BytesIO(file_content),
        read_only=True,
        data_only=True
    )

    worksheet = workbook.active

    rows = list(
        worksheet.iter_rows(
            values_only=True
        )
    )

    if not rows:
        raise ValueError(
            "Excel file is empty"
        )

    # =====================================================
    # READ HEADER
    # =====================================================

    headers = [
        str(value).strip().lower()
        if value is not None
        else ""
        for value in rows[0]
    ]

    missing_columns = (
        REQUIRED_STUDENT_COLUMNS
        - set(headers)
    )

    if missing_columns:

        raise ValueError(
            "Missing required columns: "
            + ", ".join(
                sorted(missing_columns)
            )
        )

    header_index = {
        header: index
        for index, header
        in enumerate(headers)
    }

    result = []

    # =====================================================
    # READ DATA ROWS
    # =====================================================

    for row_number, row in enumerate(
        rows[1:],
        start=2
    ):

        # ---------------------------------------------
        # Skip completely empty rows
        # ---------------------------------------------

        if not any(
            value is not None
            and str(value).strip() != ""
            for value in row
        ):
            continue

        full_name = row[
            header_index["full_name"]
        ]

        email = row[
            header_index["email"]
        ]

        class_year = row[
            header_index["class_year"]
        ]

        full_name = (
            str(full_name).strip()
            if full_name is not None
            else ""
        )

        email = (
            str(email).strip().lower()
            if email is not None
            else ""
        )

        class_year = (
            str(class_year).strip().lower()
            if class_year is not None
            else ""
        )

        result.append(
            {
                "row_number": row_number,
                "full_name": full_name,
                "email": email,
                "class_year": class_year
            }
        )

    return result

#lecturer
REQUIRED_LECTURER_COLUMNS = {
    "full_name",
    "email",
    "classes"
}


def read_lecturer_excel(
    file_content: bytes
):

    workbook = load_workbook(
        filename=io.BytesIO(file_content),
        read_only=True,
        data_only=True
    )

    worksheet = workbook.active

    rows = list(
        worksheet.iter_rows(
            values_only=True
        )
    )

    if not rows:

        raise ValueError(
            "Excel file is empty"
        )

    headers = [
        str(value).strip().lower()
        if value is not None
        else ""
        for value in rows[0]
    ]

    missing_columns = (
        REQUIRED_LECTURER_COLUMNS
        - set(headers)
    )

    if missing_columns:

        raise ValueError(
            "Missing required columns: "
            + ", ".join(
                sorted(missing_columns)
            )
        )

    header_index = {
        header: index
        for index, header
        in enumerate(headers)
    }

    result = []

    for row_number, row in enumerate(
        rows[1:],
        start=2
    ):

        full_name = row[
            header_index["full_name"]
        ]

        email = row[
            header_index["email"]
        ]

        classes = row[
            header_index["classes"]
        ]

        result.append(
            {
                "row_number": row_number,

                "full_name": (
                    str(full_name).strip()
                    if full_name is not None
                    else ""
                ),

                "email": (
                    str(email).strip().lower()
                    if email is not None
                    else ""
                ),

                "classes": (
                    str(classes).strip().lower()
                    if classes is not None
                    else ""
                )
            }
        )

    return result